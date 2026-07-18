import io
import csv
import json
from datetime import datetime
from typing import Optional
from sqlalchemy.orm import Session
from sqlalchemy import text
from app.models.evaluation import Evaluation
from app.models.lob import LOB
from app.models.pii_audit_log import PIIAuditLog


def _apply_filters(query, tenant_id: int, date_from: Optional[str] = None, date_to: Optional[str] = None, lob_id: Optional[int] = None):
    """Apply date, LOB and tenant filters to a query."""
    query = query.filter(Evaluation.tenant_id == tenant_id)
    if date_from:
        try:
            dt_from = datetime.fromisoformat(date_from)
            query = query.filter(Evaluation.evaluation_date >= dt_from)
        except ValueError:
            pass
    if date_to:
        try:
            dt_to = datetime.fromisoformat(date_to)
            dt_to = dt_to.replace(hour=23, minute=59, second=59)
            query = query.filter(Evaluation.evaluation_date <= dt_to)
        except ValueError:
            pass
    if lob_id:
        query = query.filter(Evaluation.lob_id == lob_id)
    return query


def _get_report_data(db: Session, tenant_id: int, scored_only: bool = False, date_from: str = None, date_to: str = None, lob_id: int = None):
    """Get evaluation data with optional filters."""
    query = db.query(Evaluation).order_by(Evaluation.evaluation_date.desc(), Evaluation.id.desc())
    query = _apply_filters(query, tenant_id, date_from, date_to, lob_id)
    if scored_only:
        query = query.filter(Evaluation.final_score != None)
    return query.all()


def generate_full_report(db: Session, tenant_id: int, date_from: str = None, date_to: str = None, lob_id: int = None):
    output = io.StringIO()
    writer = csv.writer(output)

    evals = _get_report_data(db, tenant_id, date_from=date_from, date_to=date_to, lob_id=lob_id)

    if not evals:
        writer.writerow(["No data found"])
        return output

    headers = [column.name for column in Evaluation.__table__.columns]
    writer.writerow(headers)

    for eval_record in evals:
        row = [getattr(eval_record, header) for header in headers]
        writer.writerow(row)

    return output


def generate_summary_report(db: Session, tenant_id: int, date_from: str = None, date_to: str = None, lob_id: int = None):
    output = io.StringIO()
    writer = csv.writer(output)

    columns = [
        "id", "call_id", "user_id", "evaluation_date", "final_score",
        "ttca_seconds", "ttch_seconds",
        "greeting", "hipaa_verification", "resolve_concern", "pci_compliance",
        "call_closing", "professionalism", "call_management", "documentation",
        "greeting_ai", "hipaa_verification_ai", "resolve_concern_ai", "pci_compliance_ai",
        "call_closing_ai", "professionalism_ai", "call_management_ai", "documentation_ai"
    ]
    writer.writerow(columns)

    evals = _get_report_data(db, tenant_id, scored_only=True, date_from=date_from, date_to=date_to, lob_id=lob_id)

    for eval_record in evals:
        row = [getattr(eval_record, col) for col in columns]
        writer.writerow(row)

    return output


def generate_ai_performance_report(db: Session, tenant_id: int, date_from: str = None, date_to: str = None, lob_id: int = None):
    output = io.StringIO()
    writer = csv.writer(output)

    headers = ['call_id', 'evaluation_date', 'criterion', 'ai_explanation', 'ai_choice', 'human_choice']
    writer.writerow(headers)

    lobs = db.query(LOB).filter(LOB.tenant_id == tenant_id).all()
    lobs_map = {lob.id: lob.criteria_json for lob in lobs}

    evals = _get_report_data(db, tenant_id, scored_only=True, date_from=date_from, date_to=date_to, lob_id=lob_id)

    for eval_record in evals:
        lob_criteria = lobs_map.get(eval_record.lob_id, {})
        for key, criteria_data in lob_criteria.items():
            if criteria_data.get("manual_score_required", False):
                continue

            human_choice = getattr(eval_record, key)
            ai_choice = getattr(eval_record, f"{key}_ai")
            reasoning = getattr(eval_record, f"{key}_reasoning", "N/A")

            if human_choice is not None and ai_choice is not None and human_choice != ai_choice:
                writer.writerow([
                    eval_record.call_id,
                    eval_record.evaluation_date,
                    criteria_data['question'],
                    reasoning,
                    ai_choice,
                    human_choice
                ])

    return output


def generate_pii_audit_report(db: Session, tenant_id: int, date_from: str = None, date_to: str = None):
    output = io.StringIO()
    writer = csv.writer(output)

    headers = ['id', 'evaluation_id', 'redacted_type', 'redacted_value_hash', 'redacted_at', 'user_id']
    writer.writerow(headers)

    query = db.query(PIIAuditLog).filter(PIIAuditLog.tenant_id == tenant_id)
    
    if date_from:
        try:
            dt_from = datetime.fromisoformat(date_from)
            query = query.filter(PIIAuditLog.redacted_at >= dt_from)
        except ValueError:
            pass
    if date_to:
        try:
            dt_to = datetime.fromisoformat(date_to)
            dt_to = dt_to.replace(hour=23, minute=59, second=59)
            query = query.filter(PIIAuditLog.redacted_at <= dt_to)
        except ValueError:
            pass
            
    audit_logs = query.order_by(PIIAuditLog.redacted_at.desc()).all()

    for log in audit_logs:
        writer.writerow([
            log.id,
            log.evaluation_id,
            log.redacted_type,
            log.redacted_value_hash,
            log.redacted_at,
            log.user_id
        ])

    return output


# ── Excel Export ──────────────────────────────────────────────────────────────

def _csv_to_excel(csv_output: io.StringIO) -> io.BytesIO:
    """Convert a CSV StringIO to an XLSX BytesIO using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # Fallback: return CSV content as bytes if openpyxl is not installed
        result = io.BytesIO()
        result.write(csv_output.getvalue().encode('utf-8'))
        return result

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    csv_output.seek(0)
    reader = csv.reader(csv_output)

    for row_idx, row in enumerate(reader, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:
                # Alternate row coloring
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    result = io.BytesIO()
    wb.save(result)
    return result


def generate_full_report_excel(db: Session, tenant_id: int, date_from: str = None, date_to: str = None, lob_id: int = None):
    csv_output = generate_full_report(db, tenant_id, date_from=date_from, date_to=date_to, lob_id=lob_id)
    return _csv_to_excel(csv_output)


def generate_summary_report_excel(db: Session, tenant_id: int, date_from: str = None, date_to: str = None, lob_id: int = None):
    csv_output = generate_summary_report(db, tenant_id, date_from=date_from, date_to=date_to, lob_id=lob_id)
    return _csv_to_excel(csv_output)


# ── PDF Export ────────────────────────────────────────────────────────────────

def _csv_to_pdf(csv_output: io.StringIO, title: str = "Report") -> io.BytesIO:
    """Convert a CSV StringIO to a PDF BytesIO using reportlab."""
    try:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.units import inch
    except ImportError:
        # Fallback: return CSV content as bytes if reportlab is not installed
        result = io.BytesIO()
        result.write(csv_output.getvalue().encode('utf-8'))
        return result

    result = io.BytesIO()
    doc = SimpleDocTemplate(result, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)

    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(f"<b>{title}</b>", styles['Title']))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Table data
    csv_output.seek(0)
    reader = csv.reader(csv_output)
    data = list(reader)

    if not data:
        elements.append(Paragraph("No data available.", styles['Normal']))
    else:
        # Limit columns for readability in PDF
        max_cols = min(len(data[0]), 12)
        trimmed_data = [row[:max_cols] for row in data]

        # Wrap long text
        wrapped_data = []
        for row_idx, row in enumerate(trimmed_data):
            wrapped_row = []
            for cell in row:
                cell_str = str(cell) if cell else ""
                if len(cell_str) > 30:
                    cell_str = cell_str[:27] + "..."
                wrapped_row.append(cell_str)
            wrapped_data.append(wrapped_row)

        table = Table(wrapped_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)

    doc.build(elements)
    return result


def generate_full_report_pdf(db: Session, tenant_id: int, date_from: str = None, date_to: str = None, lob_id: int = None):
    csv_output = generate_full_report(db, tenant_id, date_from=date_from, date_to=date_to, lob_id=lob_id)
    return _csv_to_pdf(csv_output, title="Full Database Report")


def generate_summary_report_pdf(db: Session, tenant_id: int, date_from: str = None, date_to: str = None, lob_id: int = None):
    csv_output = generate_summary_report(db, tenant_id, date_from=date_from, date_to=date_to, lob_id=lob_id)
    return _csv_to_pdf(csv_output, title="Summary Report")
